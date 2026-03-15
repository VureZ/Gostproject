# -*- coding: utf-8 -*-
"""
GOST Processor — GUI на CustomTkinter.
Запуск: python GOST_gui.py или start_gui.bat
"""

import os
import sys
import time
import threading
import customtkinter as ctk
from tkinter import messagebox

# Переменные среды для PaddleOCR
os.environ['FLAGS_use_mkldnn'] = '0'
os.environ['FLAGS_use_onednn'] = '0'
os.environ['PADDLE_USE_ONEDNN'] = '0'

# ============================================================
# Цветовая схема
# ============================================================
COLORS = {
    "bg_dark": "#1a1a2e",
    "bg_frame": "#16213e",
    "bg_card": "#0f3460",
    "accent_blue": "#0f7ddb",
    "accent_orange": "#e67e22",
    "text_white": "#ffffff",
    "text_gray": "#a0a0a0",
    "console_bg": "#0c0c0c",
    "console_fg": "#e0e0e0",
}


# ============================================================
# Общие утилиты
# ============================================================
def get_base_dir():
    """Базовая директория (рядом с EXE или рядом со скриптом)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_data_dir():
    return os.path.join(get_base_dir(), 'data')

def get_pdf_list():
    data_dir = get_data_dir()
    if not os.path.isdir(data_dir):
        return ["(нет папки data)"]
    pdfs = sorted(f for f in os.listdir(data_dir)
                  if f.endswith('.pdf') and f != 'Chat.pdf')
    return pdfs if pdfs else ["(нет PDF файлов)"]


# ============================================================
# Главное окно приложения
# ============================================================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("GOST Processor")
        self.geometry("900x650")
        self.resizable(False, False)
        self.configure(fg_color=COLORS["bg_dark"])

        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)

        self.frames = {}
        self._current_frame = None
        self._create_frames()
        self.show_frame("menu")

    def _create_frames(self):
        self.frames["menu"] = MenuFrame(self.container, self)
        self.frames["pipeline"] = PipelineFrame(self.container, self)
        self.frames["stats"] = StatsFrame(self.container, self)
        self.frames["ocr"] = OcrFrame(self.container, self)

    def show_frame(self, name: str):
        if self._current_frame:
            self._current_frame.pack_forget()
        frame = self.frames[name]
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        self._current_frame = frame

    def back_to_menu(self):
        for name, frame in self.frames.items():
            if name != "menu" and hasattr(frame, "reset"):
                frame.reset()
        self.show_frame("menu")


# ============================================================
# Экран: Главное меню
# ============================================================
class MenuFrame(ctk.CTkFrame):
    def __init__(self, parent, app: App):
        super().__init__(parent, fg_color="transparent")
        self.app = app

        ctk.CTkLabel(
            self, text="GOST Processor",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=COLORS["accent_orange"],
        ).pack(pady=(40, 5))

        ctk.CTkLabel(
            self, text="Обработка ГОСТов и генерация обозначений",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["text_gray"],
        ).pack(pady=(0, 30))

        self.desc_label = ctk.CTkLabel(
            self, text="Наведите на кнопку для описания",
            font=ctk.CTkFont(size=13),
            text_color=COLORS["text_gray"],
            wraplength=550, height=45,
        )
        self.desc_label.pack(pady=(0, 20))

        buttons_data = [
            ("📄  PDF Пайплайн", "pipeline",
             "Открыть PDF → извлечь таблицу → выбрать шаблон → сгенерировать обозначения → записать в SQL"),
            ("📊  Статистика БД", "stats",
             "Показать количество обозначений и параметров в базе данных SQL Server"),
            ("🔍  OCR Сканирование", "ocr",
             "Распознать текст из сканированного PDF (PaddleOCR) и сохранить в файл"),
        ]
        for text, target, desc in buttons_data:
            btn = ctk.CTkButton(
                self, text=text, width=340, height=50,
                font=ctk.CTkFont(size=16),
                fg_color=COLORS["accent_blue"],
                hover_color=COLORS["accent_orange"],
                corner_radius=10,
                command=lambda t=target: self.app.show_frame(t),
            )
            btn.pack(pady=6)
            btn.bind("<Enter>", lambda e, d=desc: self.desc_label.configure(text=d))
            btn.bind("<Leave>", lambda e: self.desc_label.configure(
                text="Наведите на кнопку для описания"))

        ctk.CTkButton(
            self, text="Выход", width=340, height=40,
            font=ctk.CTkFont(size=14),
            fg_color="#555555", hover_color="#cc3333",
            corner_radius=10, command=self.app.destroy,
        ).pack(pady=(20, 0))


# ============================================================
# Базовый экран операции (заголовок + назад + консоль)
# ============================================================
class BaseOperationFrame(ctk.CTkFrame):
    def __init__(self, parent, app: App, title: str):
        super().__init__(parent, fg_color="transparent")
        self.app = app
        self._process_running = False

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))

        self.back_btn = ctk.CTkButton(
            header, text="← Меню", width=100, height=32,
            font=ctk.CTkFont(size=13),
            fg_color="#555555", hover_color=COLORS["accent_blue"],
            corner_radius=8, command=self._on_back,
        )
        self.back_btn.pack(side="left")

        ctk.CTkLabel(
            header, text=title,
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=COLORS["accent_orange"],
        ).pack(side="left", padx=20)

        self.console = ctk.CTkTextbox(
            self, height=200,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color=COLORS["console_bg"],
            text_color=COLORS["console_fg"],
            corner_radius=8,
        )
        self.console.pack(fill="both", expand=True, pady=(5, 0))
        self.console.configure(state="disabled")

    def log(self, text: str):
        self.console.configure(state="normal")
        self.console.insert("end", text + "\n")
        self.console.see("end")
        self.console.configure(state="disabled")
        self.update_idletasks()

    def clear_console(self):
        self.console.configure(state="normal")
        self.console.delete("1.0", "end")
        self.console.configure(state="disabled")

    def set_process_running(self, running: bool):
        self._process_running = running
        self.back_btn.configure(state="disabled" if running else "normal")

    def _on_back(self):
        if not self._process_running:
            self.app.back_to_menu()

    def reset(self):
        self.clear_console()
        self._process_running = False
        self.back_btn.configure(state="normal")


# ============================================================
# Экран: PDF Пайплайн
# ============================================================
class PipelineFrame(BaseOperationFrame):
    def __init__(self, parent, app: App):
        super().__init__(parent, app, "PDF Пайплайн")
        self._parsed = False
        self._generated = False
        self._designations = []
        self._pipeline = None

        # --- Панель управления ---
        controls = ctk.CTkFrame(self, fg_color=COLORS["bg_frame"], corner_radius=10)
        controls.pack(before=self.console, fill="x", pady=(0, 8))

        # Строка 1: PDF + Парсить
        row1 = ctk.CTkFrame(controls, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=(10, 5))

        ctk.CTkLabel(row1, text="PDF:", font=ctk.CTkFont(size=13),
                     text_color=COLORS["text_white"]).pack(side="left")
        self.pdf_var = ctk.StringVar(value="— выберите файл —")
        self.pdf_menu = ctk.CTkOptionMenu(
            row1, variable=self.pdf_var, values=get_pdf_list(),
            width=350, font=ctk.CTkFont(size=13),
            fg_color=COLORS["bg_card"], button_color=COLORS["accent_blue"],
            button_hover_color=COLORS["accent_orange"],
        )
        self.pdf_menu.pack(side="left", padx=(8, 10))
        self.parse_btn = ctk.CTkButton(
            row1, text="▶ Парсить", width=120, height=32,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS["accent_blue"], hover_color=COLORS["accent_orange"],
            command=self._on_parse,
        )
        self.parse_btn.pack(side="left")

        # Строка 2: Шаблон
        row2 = ctk.CTkFrame(controls, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(row2, text="Шаблон:", font=ctk.CTkFont(size=13),
                     text_color=COLORS["text_white"]).pack(side="left")
        templates = [
            "1. ГОСТ 15524 стиль: М{d}-6Н.{класс}... (S{s})",
            "2. ISO стиль: ГОСТ {номер}-M{d}-{группа}",
            "3. Авто (по структуре PDF)",
        ]
        self.tpl_var = ctk.StringVar(value=templates[0])
        self.tpl_menu = ctk.CTkOptionMenu(
            row2, variable=self.tpl_var, values=templates,
            width=420, font=ctk.CTkFont(size=12),
            fg_color=COLORS["bg_card"], button_color=COLORS["accent_blue"],
            button_hover_color=COLORS["accent_orange"],
            state="disabled",
        )
        self.tpl_menu.pack(side="left", padx=(8, 0))

        # Строка 3: Материалы + кнопки действий
        row3 = ctk.CTkFrame(controls, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=(5, 10))
        ctk.CTkLabel(row3, text="Материалы:", font=ctk.CTkFont(size=13),
                     text_color=COLORS["text_white"]).pack(side="left")
        self.mat_var = ctk.StringVar(value="Стандартный набор")
        self.mat_menu = ctk.CTkOptionMenu(
            row3, variable=self.mat_var,
            values=["Стандартный набор", "Без покрытий", "Только углеродистая"],
            width=200, font=ctk.CTkFont(size=12),
            fg_color=COLORS["bg_card"], button_color=COLORS["accent_blue"],
            button_hover_color=COLORS["accent_orange"],
            state="disabled",
        )
        self.mat_menu.pack(side="left", padx=(8, 15))

        self.gen_btn = ctk.CTkButton(
            row3, text="⚙ Генерировать", width=140, height=32,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS["accent_orange"], hover_color="#d35400",
            state="disabled", command=self._on_generate,
        )
        self.gen_btn.pack(side="left", padx=(0, 8))

        self.sql_btn = ctk.CTkButton(
            row3, text="💾 В SQL", width=110, height=32,
            font=ctk.CTkFont(size=13),
            fg_color="#27ae60", hover_color="#1e8449",
            state="disabled", command=self._on_write_sql,
        )
        self.sql_btn.pack(side="left", padx=(0, 8))

        self.excel_btn = ctk.CTkButton(
            row3, text="📊 Excel", width=110, height=32,
            font=ctk.CTkFont(size=13),
            fg_color="#2980b9", hover_color="#1a5276",
            state="disabled", command=self._on_export_excel,
        )
        self.excel_btn.pack(side="left")

        self.log("Выберите PDF файл и нажмите «Парсить».")

    # --- Парсинг ---
    def _on_parse(self):
        pdf_name = self.pdf_var.get()
        if pdf_name.startswith("—") or pdf_name.startswith("("):
            self.log("⚠ Сначала выберите PDF файл!")
            return
        pdf_path = os.path.join(get_data_dir(), pdf_name)
        if not os.path.exists(pdf_path):
            self.log(f"⚠ Файл не найден: {pdf_path}")
            return
        # Сброс состояния генерации при повторном парсинге
        self._generated = False
        self._designations = []
        self.sql_btn.configure(state="disabled")
        self.excel_btn.configure(state="disabled")
        self.clear_console()
        self.log(f"Парсинг: {pdf_name}...")
        self.set_process_running(True)
        self.parse_btn.configure(state="disabled")
        threading.Thread(target=self._parse_thread, args=(pdf_path,),
                         daemon=True).start()

    def _parse_thread(self, pdf_path):
        try:
            t0 = time.time()
            # Информация о файле
            fsize = os.path.getsize(pdf_path) / 1024
            self.after(0, self.log, f"  Файл: {os.path.basename(pdf_path)} ({fsize:.0f} КБ)")

            from modules.pipeline import GostPipeline
            import fitz
            doc = fitz.open(pdf_path)
            self.after(0, self.log, f"  Страниц в PDF: {doc.page_count}")
            doc.close()

            self.after(0, self.log, "  Извлечение текста и таблиц...")
            self._pipeline = GostPipeline()
            data = self._pipeline.parser.parse_pdf(pdf_path)
            self._pipeline.parsed_data = data
            elapsed = time.time() - t0

            self.after(0, self.log, f"\n  ─── Результат парсинга ───")
            self.after(0, self.log, f"  ГОСТ:       {data.gost_number}")
            self.after(0, self.log, f"  Изделие:    {data.product_name}")
            self.after(0, self.log, f"  Диаметров:  {len(data.diameters)}")
            if data.diameters:
                d_str = ", ".join(data.diameters[:12])
                if len(data.diameters) > 12:
                    d_str += "..."
                self.after(0, self.log, f"              [{d_str}]")
            self.after(0, self.log, f"  Крупных шагов: {len(data.coarse_pitches)}")
            if data.coarse_pitches:
                items = list(data.coarse_pitches.items())[:5]
                for d, p in items:
                    self.after(0, self.log, f"    d={d} → шаг {p}")
            self.after(0, self.log, f"  Мелких шагов:  {len(data.fine_pitches)}")
            if data.fine_pitches:
                for d, plist in list(data.fine_pitches.items())[:5]:
                    self.after(0, self.log, f"    d={d} → {', '.join(plist)}")
            self.after(0, self.log, f"  Размеров S:    {len(data.wrench_sizes)}")
            if data.designation_examples:
                self.after(0, self.log, f"\n  Примеры обозначений из PDF:")
                for ex in data.designation_examples[:3]:
                    self.after(0, self.log, f"    {ex}")
            self.after(0, self.log, f"\n  Время парсинга: {elapsed:.1f} сек")

            if not data.diameters:
                self.after(0, self.log, "\n⚠ Таблица размеров не найдена!")
                self.after(0, self.log, "  Возможно PDF сканированный — попробуйте OCR.")
            else:
                self.after(0, self.log, "\n✓ Парсинг завершён. Выберите шаблон и нажмите «Генерировать».")
                self._parsed = True
                self.after(0, self._enable_after_parse)
        except Exception as e:
            self.after(0, self.log, f"\n✗ Ошибка парсинга: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.parse_btn.configure(state="normal"))

    def _enable_after_parse(self):
        self.tpl_menu.configure(state="normal")
        self.mat_menu.configure(state="normal")
        self.gen_btn.configure(state="normal")

    # --- Генерация ---
    def _on_generate(self):
        if not self._parsed or not self._pipeline:
            self.log("⚠ Сначала выполните парсинг PDF!")
            return
        self.log("\nГенерация обозначений...")
        self.sql_btn.configure(state="disabled")
        self.excel_btn.configure(state="disabled")
        self.set_process_running(True)
        self.gen_btn.configure(state="disabled")
        threading.Thread(target=self._generate_thread, daemon=True).start()

    def _generate_thread(self):
        try:
            t0 = time.time()
            tpl = self.tpl_var.get()
            mat = self.mat_var.get()
            self.after(0, self.log, f"  Шаблон: {tpl}")
            self.after(0, self.log, f"  Материалы: {mat}")

            # Шаблоны
            if tpl.startswith("1."):
                tc = "{product} М{diameter}-6Н.{group}{steel_part}{coating_part} (S{s}) ГОСТ {gost}"
                tf = "{product} 2М{diameter} × {pitch}—6Н.{group}{steel_part}{coating_part} ГОСТ {gost}"
            elif tpl.startswith("2."):
                tc = "{product} ГОСТ {gost}-M{diameter}-{group}"
                tf = "{product} ГОСТ {gost}-M{diameter}x{pitch}-{group}"
            else:
                tc, tf = "", ""

            # Материалы
            mg, sg, co, ct = None, None, None, None
            if mat == "Без покрытий":
                co, ct = {}, []
            elif mat == "Только углеродистая":
                mg = {"5": "carbon", "8": "carbon", "10": "carbon"}
                sg, co, ct = {}, {}, []

            from modules.designation_generator import DesignationGenerator
            self._pipeline.spec = self._pipeline._build_spec(
                material_groups=mg, steel_grades=sg,
                coatings=co, coating_thicknesses=ct,
            )
            self._pipeline.spec.format_coarse = tc
            self._pipeline.spec.format_fine = tf

            gen = DesignationGenerator(self._pipeline.spec)
            self._designations = gen.generate_all()
            self._pipeline.designations = self._designations

            n = len(self._designations)
            elapsed = time.time() - t0
            self.after(0, self.log, f"\n  ─── Результат генерации ───")
            self.after(0, self.log, f"  Всего обозначений: {n}")
            # Статистика
            coarse_cnt = sum(1 for d in self._designations if not d['ThreadPitch'])
            fine_cnt = n - coarse_cnt
            coated_cnt = sum(1 for d in self._designations if d['Coating'])
            uncoated_cnt = n - coated_cnt
            with_grade = sum(1 for d in self._designations if d['SteelGrade'])
            self.after(0, self.log, f"    Крупный шаг: {coarse_cnt} | Мелкий шаг: {fine_cnt}")
            self.after(0, self.log, f"    Без покрытия: {uncoated_cnt} | С покрытием: {coated_cnt}")
            self.after(0, self.log, f"    С маркой стали: {with_grade}")
            if self._designations:
                self.after(0, self.log, f"\n  Примеры:")
                for d in self._designations[:3]:
                    self.after(0, self.log, f"    {d['FullDesignation']}")
                if n > 3:
                    self.after(0, self.log, "    ...")
                    self.after(0, self.log, f"    {self._designations[-1]['FullDesignation']}")
            self.after(0, self.log, f"\n  Время генерации: {elapsed:.1f} сек")
            self.after(0, self.log, "✓ Генерация завершена. Нажмите «В SQL» для записи.")
            self._generated = True
            self.after(0, lambda: self.sql_btn.configure(state="normal"))
            self.after(0, lambda: self.excel_btn.configure(state="normal"))
        except Exception as e:
            self.after(0, self.log, f"\n✗ Ошибка генерации: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.gen_btn.configure(state="normal"))

    # --- Запись в SQL ---
    def _on_write_sql(self):
        if not self._generated or not self._designations:
            self.log("⚠ Сначала выполните генерацию!")
            return
        gost = self._pipeline.parsed_data.gost_number if self._pipeline.parsed_data else "?"
        n = len(self._designations)
        ok = messagebox.askyesno(
            "Подтверждение записи",
            f"Записать {n} обозначений ГОСТ {gost} в SQL?\n\n"
            f"Существующие записи этого ГОСТ будут удалены.",
        )
        if not ok:
            self.log("  Запись отменена.")
            return
        self.log("\nЗапись в SQL Server...")
        self.set_process_running(True)
        self.sql_btn.configure(state="disabled")
        threading.Thread(target=self._write_sql_thread, daemon=True).start()

    def _write_sql_thread(self):
        try:
            t0 = time.time()
            from modules.database import GostDatabase
            from modules.config import SQL_SERVER_CONFIG
            server = SQL_SERVER_CONFIG.get('server', '?')
            database = SQL_SERVER_CONFIG.get('database', '?')
            self.after(0, self.log, f"  Сервер: {server}")
            self.after(0, self.log, f"  База:   {database}")

            self._pipeline.parameters = self._pipeline._build_parameters()
            db = GostDatabase()
            if not db.connect():
                self.after(0, self.log, "✗ Нет подключения к SQL Server!")
                return
            self.after(0, self.log, "  Подключение установлено.")

            gost = self._pipeline.parsed_data.gost_number
            dc = db.insert_designations(self._designations, clear_existing=True)
            self.after(0, self.log, f"  Обозначений записано: {dc}")
            pc = 0
            if self._pipeline.parameters:
                pc = db.insert_parameters(self._pipeline.parameters, clear_existing=True)
                self.after(0, self.log, f"  Параметров записано:  {pc}")
            db.disconnect()

            self._pipeline._save_csv()
            csv_name = f"designations_{gost.replace(' ', '_').replace('-', '_')}.csv"
            csv_path = os.path.join(get_base_dir(), 'output', csv_name)
            self.after(0, self.log, f"  CSV: {csv_path}")

            elapsed = time.time() - t0
            self.after(0, self.log, f"\n  ─── Итог ───")
            self.after(0, self.log, f"  ГОСТ {gost} | {len(self._pipeline.parsed_data.diameters)} диаметров | {dc} обозначений | {pc} параметров")
            self.after(0, self.log, f"  SQL ✓ | CSV ✓ | {elapsed:.1f} сек")
            self.after(0, self.log, "\n✓ Запись завершена!")
        except Exception as e:
            self.after(0, self.log, f"\n✗ Ошибка SQL: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.sql_btn.configure(state="normal"))

    # --- Экспорт в Excel ---
    def _on_export_excel(self):
        if not self._generated or not self._designations:
            self.log("⚠ Сначала выполните генерацию!")
            return
        dialog = ctk.CTkInputDialog(
            text="Введите название файла (без .xlsx):\n(по умолчанию: GOST)",
            title="Экспорт в Excel",
        )
        filename = dialog.get_input()
        if filename is None:
            self.log("  Экспорт отменён.")
            return
        filename = filename.strip() if filename else ""
        if not filename:
            filename = "GOST"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        out_dir = os.path.join(get_base_dir(), 'output')
        os.makedirs(out_dir, exist_ok=True)
        filepath = os.path.join(out_dir, filename)
        if os.path.exists(filepath):
            ok = messagebox.askyesno(
                "Файл уже существует",
                f"Файл «{filename}» уже существует.\nЗаменить?",
            )
            if not ok:
                self.log("  Экспорт отменён.")
                return
        self.log(f"\nЭкспорт в Excel: {filename}...")
        self.set_process_running(True)
        self.excel_btn.configure(state="disabled")
        threading.Thread(target=self._excel_thread,
                         args=(filepath, filename), daemon=True).start()

    def _excel_thread(self, filepath, filename):
        try:
            t0 = time.time()
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = Workbook()
            ws = wb.active
            gost = self._pipeline.parsed_data.gost_number if self._pipeline.parsed_data else "GOST"
            ws.title = f"ГОСТ {gost}"[:31]  # Excel ограничение 31 символ
            # Заголовки
            headers = ["№", "ГОСТ", "Обозначение", "Резьба",
                       "Группа", "Покрытие", "Марка стали", "Диаметр", "Шаг"]
            hdr_font = Font(bold=True, size=11, color="FFFFFF")
            hdr_fill = PatternFill(start_color="0f7ddb", end_color="0f7ddb", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side('thin'), right=Side('thin'),
                            top=Side('thin'), bottom=Side('thin'))
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
                c.border = border
            # Данные
            keys = ['GOST_Number', 'FullDesignation', 'ThreadSize',
                    'MaterialGroup', 'Coating', 'SteelGrade',
                    'ThreadDiameter', 'ThreadPitch']
            for i, d in enumerate(self._designations, 1):
                ws.cell(row=i+1, column=1, value=i).border = border
                for j, k in enumerate(keys, 2):
                    ws.cell(row=i+1, column=j, value=d.get(k, '')).border = border
            # Авто-ширина
            widths = [6, 18, 55, 14, 10, 12, 14, 10, 8]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + i)].width = w
            ws.freeze_panes = "A2"
            wb.save(filepath)
            elapsed = time.time() - t0
            n = len(self._designations)
            self.after(0, self.log, f"  Записей: {n}")
            self.after(0, self.log, f"  Файл: {filepath}")
            self.after(0, self.log, f"  Время: {elapsed:.1f} сек")
            self.after(0, self.log, "\n✓ Excel экспорт завершён!")
        except Exception as e:
            self.after(0, self.log, f"\n✗ Ошибка Excel: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.excel_btn.configure(state="normal"))

    # --- Сброс ---
    def reset(self):
        super().reset()
        self._parsed = False
        self._generated = False
        self._designations = []
        self._pipeline = None
        self.pdf_var.set("— выберите файл —")
        self.pdf_menu.configure(values=get_pdf_list())
        self.tpl_var.set("1. ГОСТ 15524 стиль: М{d}-6Н.{класс}... (S{s})")
        self.mat_var.set("Стандартный набор")
        self.tpl_menu.configure(state="disabled")
        self.mat_menu.configure(state="disabled")
        self.gen_btn.configure(state="disabled")
        self.sql_btn.configure(state="disabled")
        self.excel_btn.configure(state="disabled")
        self.parse_btn.configure(state="normal")
        self.log("Выберите PDF файл и нажмите «Парсить».")


# ============================================================
# Экран: Статистика БД
# ============================================================
class StatsFrame(BaseOperationFrame):
    def __init__(self, parent, app: App):
        super().__init__(parent, app, "Статистика БД")
        self.load_btn = ctk.CTkButton(
            self, text="Загрузить статистику", width=250, height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["accent_blue"],
            hover_color=COLORS["accent_orange"],
            corner_radius=8, command=self._load_stats,
        )
        self.load_btn.pack(before=self.console, pady=(0, 10))

    def _load_stats(self):
        self.clear_console()
        self.set_process_running(True)
        self.load_btn.configure(state="disabled")
        self.log("Подключение к SQL Server...")
        threading.Thread(target=self._load_stats_thread, daemon=True).start()

    def _load_stats_thread(self):
        try:
            t0 = time.time()
            from modules.database import GostDatabase
            from modules.config import SQL_SERVER_CONFIG
            server = SQL_SERVER_CONFIG.get('server', '?')
            database = SQL_SERVER_CONFIG.get('database', '?')
            self.after(0, self.log, f"  Сервер: {server}")
            self.after(0, self.log, f"  База:   {database}")

            db = GostDatabase()
            if not db.connect():
                self.after(0, self.log, "ОШИБКА: нет подключения к SQL Server!")
                return
            self.after(0, self.log, "  Подключение установлено.\n")

            total = db.get_designation_count()
            self.after(0, self.log, f"  ─── Обозначения ───")
            self.after(0, self.log, f"  Всего: {total}")
            db.cursor.execute(
                "SELECT GOST_Number, COUNT(*) FROM ProductDesignations "
                "GROUP BY GOST_Number ORDER BY GOST_Number")
            rows = db.cursor.fetchall()
            if rows:
                self.after(0, self.log, "")
                self.after(0, self.log, f"  {'ГОСТ':<25} {'Обозначений':>12}")
                self.after(0, self.log, f"  {'─'*25} {'─'*12}")
                for r in rows:
                    self.after(0, self.log, f"  {r[0]:<25} {r[1]:>12}")
                self.after(0, self.log, f"  {'─'*25} {'─'*12}")
                self.after(0, self.log, f"  {'ИТОГО':<25} {total:>12}")

            db.cursor.execute("SELECT COUNT(*) FROM ProductParameters")
            p_total = db.cursor.fetchone()[0]
            self.after(0, self.log, f"\n  ─── Параметры ───")
            self.after(0, self.log, f"  Всего: {p_total}")

            if p_total > 0:
                db.cursor.execute(
                    "SELECT GOST_Number, COUNT(*) FROM ProductParameters "
                    "GROUP BY GOST_Number ORDER BY GOST_Number")
                for r in db.cursor.fetchall():
                    self.after(0, self.log, f"    {r[0]}: {r[1]}")

            db.cursor.execute(
                "SELECT TOP 10 FullDesignation FROM ProductDesignations ORDER BY ID DESC")
            rows = db.cursor.fetchall()
            if rows:
                self.after(0, self.log, f"\n  ─── Последние 10 записей ───")
                for i, r in enumerate(rows, 1):
                    self.after(0, self.log, f"  {i:2d}. {r[0]}")

            elapsed = time.time() - t0
            db.disconnect()
            self.after(0, self.log, f"\n✓ Загрузка завершена ({elapsed:.1f} сек)")
        except Exception as e:
            self.after(0, self.log, f"Ошибка: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.load_btn.configure(state="normal"))

    def reset(self):
        super().reset()
        self.load_btn.configure(state="normal")


# ============================================================
# Экран: OCR Сканирование
# ============================================================
class OcrFrame(BaseOperationFrame):
    def __init__(self, parent, app: App):
        super().__init__(parent, app, "OCR Сканирование")

        controls = ctk.CTkFrame(self, fg_color=COLORS["bg_frame"], corner_radius=10)
        controls.pack(before=self.console, fill="x", pady=(0, 8))
        row = ctk.CTkFrame(controls, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(row, text="PDF:", font=ctk.CTkFont(size=13),
                     text_color=COLORS["text_white"]).pack(side="left")
        self.pdf_var = ctk.StringVar(value="— выберите файл —")
        self.pdf_menu = ctk.CTkOptionMenu(
            row, variable=self.pdf_var, values=get_pdf_list(),
            width=350, font=ctk.CTkFont(size=13),
            fg_color=COLORS["bg_card"], button_color=COLORS["accent_blue"],
            button_hover_color=COLORS["accent_orange"],
        )
        self.pdf_menu.pack(side="left", padx=(8, 10))
        self.ocr_btn = ctk.CTkButton(
            row, text="🔍 Распознать", width=150, height=32,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS["accent_blue"], hover_color=COLORS["accent_orange"],
            command=self._on_ocr,
        )
        self.ocr_btn.pack(side="left")
        self.log("Выберите сканированный PDF и нажмите «Распознать».")
        self.log("Результат будет сохранён в папку output/.")

    def _on_ocr(self):
        pdf_name = self.pdf_var.get()
        if pdf_name.startswith("—") or pdf_name.startswith("("):
            self.log("⚠ Сначала выберите PDF файл!")
            return
        pdf_path = os.path.join(get_data_dir(), pdf_name)
        if not os.path.exists(pdf_path):
            self.log(f"⚠ Файл не найден: {pdf_path}")
            return
        self.clear_console()
        fsize = os.path.getsize(pdf_path) / 1024
        self.log(f"OCR: {pdf_name} ({fsize:.0f} КБ)")
        self.log(f"Это может занять 20-60 секунд...")
        self.set_process_running(True)
        self.ocr_btn.configure(state="disabled")
        threading.Thread(target=self._ocr_thread, args=(pdf_path, pdf_name),
                         daemon=True).start()

    def _ocr_thread(self, pdf_path, pdf_name):
        try:
            from modules.pdf_processor import PDFProcessor
            from modules.ocr_engine import OCREngine
            from modules.table_parser import TableExtractor
            import numpy as np
        except Exception as e:
            msg = str(e)
            if getattr(sys, 'frozen', False):
                self.after(0, self.log, "✗ OCR недоступен в EXE-версии.")
                self.after(0, self.log, "  PaddleOCR требует запуска через start_gui.bat")
                self.after(0, self.log, f"  (причина: {msg[:80]})")
            else:
                self.after(0, self.log, f"✗ Ошибка загрузки OCR: {msg}")
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.ocr_btn.configure(state="normal"))
            return

        try:
            t0 = time.time()
            self.after(0, self.log, "\n[1/3] Конвертация PDF в изображения...")
            images = PDFProcessor(pdf_path).convert_to_images(dpi=300)
            n_pages = len(images)
            self.after(0, self.log, f"  Страниц: {n_pages}")

            self.after(0, self.log, "\n[2/3] Загрузка OCR движка (PaddleOCR)...")
            ocr = OCREngine(language='ru', use_gpu=False)
            ext = TableExtractor(ocr_engine=ocr)
            self.after(0, self.log, "  Движок загружен.")

            self.after(0, self.log, f"\n[3/3] Распознавание текста...")
            pages = []
            for i, img in enumerate(images):
                pct = int((i + 1) / n_pages * 100)
                self.after(0, self.log, f"\n  Стр. {i+1}/{n_pages} ({pct}%)...")
                tp = time.time()
                try:
                    text = ext.get_full_page_text(np.array(img))
                    pages.append(text)
                    page_time = time.time() - tp
                    self.after(0, self.log, f"    {len(text)} символов ({page_time:.1f} сек)")
                    # Превью первых 150 символов
                    if text.strip():
                        preview = text.strip()[:150].replace('\n', ' ')
                        self.after(0, self.log, f"    Превью: {preview}...")
                except Exception as e:
                    pages.append("")
                    self.after(0, self.log, f"    Ошибка: {e}")

            # Сохранение
            out_dir = os.path.join(get_base_dir(), 'output')
            os.makedirs(out_dir, exist_ok=True)
            fn = os.path.splitext(pdf_name)[0]
            out_path = os.path.join(out_dir, f'ocr_{fn}.txt')
            with open(out_path, 'w', encoding='utf-8') as f:
                for i, t in enumerate(pages):
                    f.write(f"{'='*60}\nPAGE {i+1}\n{'='*60}\n{t}\n\n")

            total_chars = sum(len(p) for p in pages)
            elapsed = time.time() - t0
            self.after(0, self.log, f"\n  ─── Итог ───")
            self.after(0, self.log, f"  Страниц: {n_pages}")
            self.after(0, self.log, f"  Символов: {total_chars}")
            self.after(0, self.log, f"  Время: {elapsed:.1f} сек")
            self.after(0, self.log, f"  Файл: {out_path}")
            self.after(0, self.log, "\n✓ OCR завершён!")
        except Exception as e:
            self.after(0, self.log, f"\n✗ Ошибка OCR: {e}")
        finally:
            self.after(0, self.set_process_running, False)
            self.after(0, lambda: self.ocr_btn.configure(state="normal"))

    def reset(self):
        super().reset()
        self.pdf_var.set("— выберите файл —")
        self.pdf_menu.configure(values=get_pdf_list())
        self.ocr_btn.configure(state="normal")
        self.log("Выберите сканированный PDF и нажмите «Распознать».")
        self.log("Результат будет сохранён в папку output/.")


# ============================================================
# Точка входа
# ============================================================
if __name__ == "__main__":
    try:
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        app = App()
        app.mainloop()
    except Exception as e:
        import tkinter.messagebox as mb
        mb.showerror("GOST Processor — Ошибка", str(e))
